from __future__ import annotations

import csv
import uuid
import logging
from pathlib import Path

import config

logger = logging.getLogger(__name__)


def save_upload(data: bytes, filename: str) -> Path:
    """Save uploaded file to workspace/uploads. Returns the file path."""
    if len(data) > config.MAX_FILE_SIZE_BYTES:
        raise ValueError(f"File too large: {len(data)} bytes (max {config.MAX_FILE_SIZE_MB}MB)")

    # Sanitize filename: strip path components to prevent traversal
    safe_name = Path(filename).name
    if not safe_name or safe_name.startswith("."):
        safe_name = f"upload_{safe_name}" if safe_name else "upload"

    # UUID suffix guarantees uniqueness without TOCTOU race conditions
    stem = Path(safe_name).stem
    suffix = Path(safe_name).suffix
    unique_name = f"{stem}_{uuid.uuid4().hex[:8]}{suffix}"
    dest = config.UPLOADS_DIR / unique_name

    dest.write_bytes(data)
    logger.info("Saved upload: %s (%d bytes)", dest.name, len(data))
    return dest


def get_file_content(path: Path, max_chars: int = 50000) -> str:
    """Read file content as text, truncating if needed."""
    try:
        text = path.read_text(encoding="utf-8", errors="replace")
        if len(text) > max_chars:
            return text[:max_chars] + f"\n... (truncated, {len(text)} total chars)"
        return text
    except Exception:
        try:
            size = path.stat().st_size
        except OSError:
            size = 0
        return f"[Binary file: {path.name}, {size} bytes]"


def get_file_metadata(path: Path) -> dict:
    """Extract metadata from a data file WITHOUT loading it into memory.

    Returns dict with: name, size_bytes, size_human, type, columns, row_count,
    sample_rows, sheet_names (Excel only).
    """
    stat = path.stat()
    size = stat.st_size
    size_human = f"{size / 1024:.1f}KB" if size < 1_000_000 else f"{size / 1_048_576:.1f}MB"

    meta = {
        "name": path.name,
        "size_bytes": size,
        "size_human": size_human,
        "type": path.suffix.lstrip("."),
        "columns": [],
        "row_count": 0,
        "sample_rows": [],
        "sheet_names": [],
    }

    try:
        if path.suffix in (".csv", ".tsv"):
            sep = "\t" if path.suffix == ".tsv" else ","
            with open(path, "r", encoding="utf-8", errors="replace") as f:
                reader = csv.reader(f, delimiter=sep)
                header = next(reader, [])
                sample = []
                data_row_count = 0
                for i, row in enumerate(reader):
                    if i < 5:
                        sample.append(row)
                    data_row_count = i + 1
            row_count = data_row_count
            meta["columns"] = header
            meta["row_count"] = row_count
            meta["sample_rows"] = sample

        elif path.suffix == ".xlsx":
            try:
                import openpyxl
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                meta["sheet_names"] = wb.sheetnames
                # Get first sheet info
                ws = wb[wb.sheetnames[0]]
                rows = list(ws.iter_rows(max_row=6, values_only=True))
                if rows:
                    meta["columns"] = [str(c) if c else "" for c in rows[0]]
                    meta["sample_rows"] = [[str(c) if c else "" for c in r] for r in rows[1:]]
                meta["row_count"] = ws.max_row or 0
                wb.close()
            except ImportError:
                meta["columns"] = ["(openpyxl not installed — use pandas.read_excel())"]

        elif path.suffix == ".json":
            import json
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, list) and len(data) > 0 and isinstance(data[0], dict):
                meta["columns"] = list(data[0].keys())
                meta["row_count"] = len(data)
                meta["sample_rows"] = [list(d.values()) for d in data[:5]]
            elif isinstance(data, dict):
                meta["columns"] = list(data.keys())

        elif path.suffix == ".parquet":
            try:
                import pyarrow.parquet as pq
                pf = pq.ParquetFile(path)
                schema = pf.schema_arrow
                meta["columns"] = [f.name for f in schema]
                meta["row_count"] = pf.metadata.num_rows
            except ImportError:
                meta["columns"] = ["(pyarrow not installed)"]

    except Exception as e:
        logger.warning("Failed to extract metadata from %s: %s", path.name, e)

    return meta


def format_file_metadata_for_prompt(path: Path) -> str:
    """Format file metadata as a prompt-friendly string for Claude.

    Does NOT load raw data into the prompt — only metadata + sample.
    """
    meta = get_file_metadata(path)

    parts = [f"--- File: {meta['name']} ({meta['size_human']}"]
    if meta["row_count"]:
        parts[0] += f", ~{meta['row_count']:,} data rows"
    parts[0] += ") ---"

    if meta["sheet_names"]:
        parts.append(f"Sheets: {meta['sheet_names']}")
    if meta["columns"]:
        parts.append(f"Columns: {meta['columns']}")
    if meta["sample_rows"]:
        parts.append(f"Sample (first {len(meta['sample_rows'])} rows): {meta['sample_rows']}")

    parts.append("DO NOT load this file into context. Write a script to process it locally.")

    return "\n".join(parts)


